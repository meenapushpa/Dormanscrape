# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html
import mysql.connector
from mysql.connector.errors import Error
from openpyxl import load_workbook, Workbook
from .. import models
from scrapy import signals
from scrapy.exporters import CsvItemExporter


class DormanproductPipeline(object):
   
   # @classmethod
   # def from_crawler(cls, crawler):
   #     settings = crawler.settings
    #    stop_flag = settings.get('STOP_FLAG')
     #   print(stop_flag)
      #  return cls(stop_flag)

    #def __init__(self):
        #self.stop_flag = stop_flag
        #self.create_connection()
        
  #  def create_connection(self):
   #     self.conn = mysql.connector.connect(
    #        host='localhost',
     ##       user='dds',     
     #         password='aw3SDer45',
      #      database='dds_web',
       #     port=3306
       # )
        #self.curr = self.conn.cursor()

    def process_item(self, item, spider):
        if models.dorman.objects.filter(original_part_number=item['partno'], product_name=item['ProductName'], part_no=item['OEref'], manufacturer_name=item['MfrName']).exists() is False:
            models.dorman.objects.create(original_part_number=item['partno'], product_name=item['ProductName'], product_info=item['ProductInfo'],application_summary=item['ProductSummary'], cross=item['Cross'], part_no=item['OEref'], manufacturer_name=item['MfrName'])
        #self.store_db(item)
        #if self.stop_flag != True:
        #else:
         #   spider.close_down = True
        return item
    
    #def store_db(self, item):
     #   try:
      #      self.curr.execute("""insert into scraper_dorman_1 values (%s, %s, %s, %s,%s,%s,%s,%s)""", (item['partno'],item['ProductName'],item['ProductInfo'],item['ProductSummary'],item['Cross'],item['OEref'],item['MfrName'],item['Oedetails']))
       #     self.conn.commit()
        #    print(self.curr.rowcount,"row inserted")
#
 #       except mysql.connector.Error as e:
  ##          print("Error code: %d,Error msg: %s",e.errno,e.msg)
       
    #def close_spider(self, spider):
     #   self.curr.close()
      #  self.conn.close()

#log runs into back file 
class JsonWriterPipeline(object):

    @classmethod
    def from_crawler(cls, crawler):
        return cls(crawler.settings)

    def __init__(self,settings):
        self.filename=settings.get('DOWNLOAD_DIR')
        #self.stop_flag=settings.get('STOP_FLAG')

    def process_item(self, item, spider):
        data=list(item.values())
        self.append_df_to_excel4(self.filename + ".xlsx",data)
        return item
       
     
    def append_df_to_excel4(self,filename,data):
        try:
            book = load_workbook(filename)
            sheet = book.active
        except:
            book = Workbook()
            sheet = book.active
            headers = ['original part number', 'Product Name', 'Product Info',
                       'Application Summary', 'Cross', 'Part No', 'Mfr Name','Oedetails']
            sheet.append(headers)

        sheet.append(data)
        book.save(filename)

class InMemoryItemStore(object):
    __ITEM_STORE = None

    @classmethod
    def pop_items(cls):
        items = cls.__ITEM_STORE or []
        cls.__ITEM_STORE = None
        return items

    @classmethod
    def add_item(cls, item):
        if not cls.__ITEM_STORE:
            cls.__ITEM_STORE = []
        cls.__ITEM_STORE.append(item)

    def process_item(self, item, spider):
        print('add item process in pipelines')
        self.add_item(item)
        return item

class CSVPipeline(object):
    @classmethod
    def from_crawler(cls, crawler):
        pipeline = cls(crawler.settings)
        crawler.signals.connect(pipeline.spider_opened, signals.spider_opened)
        crawler.signals.connect(pipeline.spider_closed, signals.spider_closed)
        return pipeline
    
    def __init__(self,pipeline):
        self.filename=pipeline.get('DOWNLOAD_DIR')

    def spider_opened(self, spider):
        self.file = open(self.filename + ".csv", 'w+b')
        self.exporter = CsvItemExporter(self.file)
        self.exporter.start_exporting()

    def spider_closed(self, spider):
        self.exporter.finish_exporting()
        self.file.close()

    def process_item(self, item, spider):
        self.exporter.export_item(item)
        return item

