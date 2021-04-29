from bs4 import BeautifulSoup
from django.core.files.storage import FileSystemStorage
from django.http import JsonResponse
from django.shortcuts import render
from django.views.generic import DetailView
from lxml.html import fromstring
import logging
from openpyxl import load_workbook, Workbook
from pyvirtualdisplay import Display
from requests.adapters import HTTPAdapter
from urllib.parse import quote_plus
from selenium import webdriver
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from sys import platform
from .dormanproject.crawl2 import crawl
from . import freeproxy
import pandas as pd
import json
import multiprocessing as mp
import copy
import glob
import json
import os
import pandas as pd
import random
import requests
import time

from . import models
BASE_CRAIGSLIST_URL = 'https://losangeles.craigslist.org/search/?query={}'
BASE_IMAGE_URL = 'https://images.craigslist.org/{}_300x300.jpg'

BASE_TARGET_URL1 = 'https://fme-cat.com/PartInterchange.aspx?pn={}'
BASE_TARGET_URL2 = 'http://www.fme-cat.mx/PartInterchange.aspx?pn={}'
BASE_TARGET_URL3 = 'https://partsonline.mevotech.com/api/search/interchange/{}?&lang=en'
BASE_TARGET_URL4 = 'https://www.dormanproducts.com/gsearch.aspx?type=oesearch&origin=oesearch&q={}'
BASE_TARGET_URL5 = 'https://nexcat.com'
BASE_TARGET_URL6 = 'https://showmetheparts.com/trw/'

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.130 Safari/537.36',
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9"
}

username = password = ''
nexcat_login_file = "nexcat_login.info"
headers5 = {
    'X-Requested-With': 'XMLHttpRequest',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.75 Safari/537.36',
    'Sec-Fetch-Mode': 'cors',
    'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
    'Referer': 'https://www.nexcat.com/nexpartbg.php',
    'Accept-Encoding': 'gzip, deflate, br',
    'Accept-Language': 'cs-CZ,cs;q=0.9,sk;q=0.8,en;q=0.7',
    'Cookie': 'memberme=on; default_order_type_id=3000483; autosort_default_branch=adminOff; includeHeavyDuty=1; PHPSESSID=lo0gmq08qbggf2i2erhughd164; login=DELPHI6; pword=RGVsVGVjNTAx; bi_log_sc_header_id=H20191001472138411454314; bi_log_sc_group_id=G20191001472138411454314'
}

proxy_list = []
with open("proxy.lst", 'r') as fp:
    for line in fp:
        proxy_list.append(line.strip())

proxies = copy.deepcopy(proxy_list)

json_file = open('headers4.json')
headers4 = json.load(json_file)
json_file.close()

flag_stop = False
server_status = ""


class ScraperView(DetailView):
    # template_name = 'base.html'
    template_name = 'scraper/show_result.html'

    def get(self, request, *args, **kwargs):
        global flag_stop
        global username
        global password

        flag_stop = False

        files = list(filter(os.path.isfile, glob.glob("static/result/*.xlsx")))
        files.sort(key=lambda x: os.path.getatime(x), reverse=True)
        files = [os.path.basename(file) for file in files]

        with open(nexcat_login_file, 'r') as fp:
            for line in fp:
                infos = line.split(':', 1)

                if infos[0].strip() == "username":
                    username = infos[1].strip()

                if infos[0].strip() == "password":
                    password = infos[1].strip()

        return render(request, self.template_name, {
            'username': username,
            'password': password,
            'search': '',
            'files': files,
        })

    def post(self, request):
        global flag_stop
        global server_status
        flag_stop = False
        server_status = ''
        new_search = False
        final_postings = []

        list1 = []
        if request.POST.get('ckb_site1') == 'on' and flag_stop is False:
            if request.POST.get('ckb_db') == 'on':
                list1 = search_db1(request)
            else:
                list1 = search1(request)
        final_postings.append(list1)

        list2 = []
        if request.POST.get('ckb_site2') == 'on' and flag_stop is False:
            if request.POST.get('ckb_db') == 'on':
                list2 = search_db2(request)
            else:
                list2 = search2(request)
        final_postings.append(list2)

        list3 = []
        if request.POST.get('ckb_site3') == 'on' and flag_stop is False:
            if request.POST.get('ckb_db') == 'on':
                list3 = search_db3(request)
            else:
                list3 = search3(request)
        final_postings.append(list3)

        list4 = []
        if request.POST.get('ckb_site4') == 'on' and flag_stop is False:
            if request.POST.get('ckb_db') == 'on':
                list4 = search_db4(request)
            elif request.method == 'POST' and request.FILES and flag_stop is False:
                list4 = scrapesearch(request)
                #list4 = new_search4(request)
            else: 
                list4 = search4(request)
        final_postings.append(list4)

        list5 = []
        if request.POST.get('ckb_site5') == 'on' and flag_stop is False:
            if request.POST.get('ckb_db') == 'on':
                if request.method == 'POST' and request.FILES:
                    list5 = new_search_db5(request)
                    new_search = True
                else:
                    list5 = search_db5(request)
            else:
                if request.method == 'POST' and request.FILES:
                    list5 = new_search5(request)
                    new_search = True
                else:
                    list5 = search5(request)
        final_postings.append(list5)

        list6 = []
        if request.POST.get('ckb_site6') == 'on' and flag_stop is False:
            if request.POST.get('ckb_db') == 'on':
                list6 = search_db6(request)
            else:
                list6 = search6(request)
        final_postings.append(list6)

        files = list(filter(os.path.isfile, glob.glob("static/result/*.xlsx")))
        files.sort(key=lambda x: os.path.getatime(x), reverse=True)
        files = [os.path.basename(file) for file in files]

        stuff_for_frontend = {
            'username': username,
            'password': password,
            'search': request.POST.get('search'),
            'files': files,
            'final_postings': final_postings,
            'new_search': new_search
        }
        return render(request, 'scraper/show_result.html', stuff_for_frontend)


def get_status(request):
    global server_status
    status = {'status': server_status[1:]}
    server_status = ""
    return JsonResponse(status)


def stop_search(request):
    global flag_stop
    flag_stop = True

    return JsonResponse({'status': 'success'})


def save_userinfo(request):
    global username
    global password
    username = request.POST.get('username')
    password = request.POST.get('password')
    f = open(nexcat_login_file, "w")
    f.write("username: " + username + '\n')
    f.write("password: " + password + '\n')
    f.close()

    return JsonResponse({'status': 'success'})


def search1(request):
    output_filename = "static/result/fme_cat_output"
    if request.POST.get('ckb_autonew') == 'on':
        fileIndex = 1
        fname = output_filename
        while os.path.isfile(fname + ".xlsx") == True:
            fname = "%s_%06d" % (output_filename, fileIndex)
            fileIndex = fileIndex + 1
        output_filename = fname

    search = request.POST.get('search')
    part_numbers = search.split()

    global server_status
    final_postings = []
    for part in part_numbers:
        final_url = BASE_TARGET_URL1.format(quote_plus(part))
        response = requests.get(final_url, headers=headers)
        print(response.status_code, response.url)
        server_status = server_status + "\n" + "{}, {}".format(response.status_code, response.url)

        html = fromstring(response.content)

        for row in html.xpath('//table[@id="MainPH_MainPH_grdIntPart"]/tr[@data-type="int"]'):
            part_info = []

            try:
                Mfr_Part = row.xpath(
                    './/*[contains(@id, "MainPH_MainPH_grdIntPart_lblMfrPart")]/text()')[0].strip()
                part_info.append(Mfr_Part)
            except:
                pass

            try:
                Our_Part = row.xpath(
                    './/*[contains(@id, "MainPH_MainPH_grdIntPart_hlOurPart")]/text()')[0].strip()
                part_info.append(Our_Part)
            except:
                pass

            try:
                mfr_name = row.xpath(
                    './/*[contains(@id, "MainPH_MainPH_grdIntPart_lblMfrName")]/text()')[0].strip()
                part_info.append(mfr_name)
            except:
                pass

            try:
                Product_Type = row.xpath(
                    './/*[contains(@id, "MainPH_MainPH_grdIntPart_lblProdType")]/text()')[0].strip()
                part_info.append(Product_Type)
            except:
                pass

            final_postings.append(part_info)
            append_df_to_excel(output_filename + ".xlsx", part_info)

            if models.fme_cat.objects.filter(original_part_number=part_info[0], part_no=part_info[1], manufacturer_name=part_info[2]).exists() is False:
                models.fme_cat.objects.create(
                    original_part_number=part_info[0], part_no=part_info[1], manufacturer_name=part_info[2], product_type=part_info[3])

        if flag_stop == True:
            break

    return final_postings


def search2(request):
    output_filename = "static/result/fme_mx_output"
    if request.POST.get('ckb_autonew') == 'on':
        fileIndex = 1
        fname = output_filename
        while os.path.isfile(fname + ".xlsx") == True:
            fname = "%s_%06d" % (output_filename, fileIndex)
            fileIndex = fileIndex + 1
        output_filename = fname

    search = request.POST.get('search')
    part_numbers = search.split()

    global server_status
    final_postings = []
    for part in part_numbers:
        final_url = BASE_TARGET_URL2.format(quote_plus(part))
        response = requests.get(final_url, headers=headers)
        print(response.status_code, response.url)
        server_status = server_status + "\n" + "{}, {}".format(response.status_code, response.url)

        soup = BeautifulSoup(response.content, 'html.parser')
        rows = soup.select('#MainPH_MainPH_grdIntPart .partRow')

        for row in rows:
            MfrName = row.select_one('span[id^="MainPH_MainPH_grdIntPart_lblMfrName_"]').text
            ProdType = row.select_one('span[id^="MainPH_MainPH_grdIntPart_lblProdType_"]').text
            MfrPart = row.select_one('span[id^="MainPH_MainPH_grdIntPart_lblMfrPart_"]').text
            hlOurPart = row.select_one('a[id^="MainPH_MainPH_grdIntPart_hlOurPart_"]').text

            final_postings.append([MfrPart, hlOurPart, MfrName, ProdType])
            append_df_to_excel(output_filename + ".xlsx", [MfrPart, hlOurPart, MfrName, ProdType])

            if models.fme_mx.objects.filter(original_part_number=MfrPart, part_no=hlOurPart, manufacturer_name=MfrName).exists() is False:
                models.fme_mx.objects.create(
                    original_part_number=MfrPart, part_no=hlOurPart, manufacturer_name=MfrName, product_type=ProdType)

        if flag_stop == True:
            break

    return final_postings


def search3(request):
    output_filename = "static/result/mevotech_output"
    if request.POST.get('ckb_autonew') == 'on':
        fileIndex = 1
        fname = output_filename
        while os.path.isfile(fname + ".xlsx") == True:
            fname = "%s_%06d" % (output_filename, fileIndex)
            fileIndex = fileIndex + 1
        output_filename = fname

    search = request.POST.get('search')
    part_numbers = search.split()

    global server_status
    final_postings = []
    for part in part_numbers:
        final_url = BASE_TARGET_URL3.format(quote_plus(part))
        response = requests.get(final_url, headers=headers)
        print(response.status_code, response.url)
        server_status = server_status + "\n" + "{}, {}".format(response.status_code, response.url)

        results = response.json()

        if 'status' in results:
            if results['status'] == 'redirect':
                part_number_url = f"https://partsonline.mevotech.com/api/part/{results['id']}?captcha=null&lang=en"

                # , verify=False, proxies=proxy)
                response = requests.get(part_number_url, headers=headers)
                print(response.status_code, response.url)
                server_status = server_status + "\n" + \
                    "{}, {}".format(response.status_code, response.url)

                product = response.json()

                final_postings.append([part.strip(), product['info']['partNo'], product['info']
                                       ['brand']['brandName'], product['info']['productDescr']])
                append_df_to_excel(output_filename + ".xlsx", [part.strip(
                ), product['info']['partNo'], product['info']['brand']['brandName'], product['info']['productDescr']])

                if models.mevotech.objects.filter(original_part_number=part.strip(), part_no=product['info']['partNo'], manufacturer_name=product['info']['brand']['brandName']).exists() is False:
                    models.mevotech.objects.create(original_part_number=part.strip(
                    ), part_no=product['info']['partNo'], manufacturer_name=product['info']['brand']['brandName'], product_type=product['info']['productDescr'])
        else:
            for partCategoryListing in results[0]['partCategoryListing']:
                for brand in partCategoryListing['brandData']:
                    for item in brand['partListing']:
                        final_postings.append([part.strip(), item['partNo'],
                                               brand['brandName'], item['productDescr']])
                        append_df_to_excel(
                            output_filename + ".xlsx", [part.strip(), item['partNo'], brand['brandName'], item['productDescr']])

                        if models.mevotech.objects.filter(original_part_number=part.strip(), part_no=item['partNo'], manufacturer_name=brand['brandName']).exists() is False:
                            models.mevotech.objects.create(original_part_number=part.strip(
                            ), part_no=item['partNo'], manufacturer_name=brand['brandName'], product_type=item['productDescr'])

        if flag_stop == True:
            break

    return final_postings


def search4(request):
    output_filename = "static/result/dorman_output"
    if request.POST.get('ckb_autonew') == 'on':
        fileIndex = 1
        fname = output_filename
        while os.path.isfile(fname + ".xlsx") == True:
            fname = "%s_%06d" % (output_filename, fileIndex)
            fileIndex = fileIndex + 1
        output_filename = fname

    search = request.POST.get('search')
    part_numbers = search.split()

    global proxy_index
    global part_index

    final_postings = []
    for part in part_numbers:
        part_index = part_index + 1
        if part_index % 15 == 0:
            part_index = 1
            proxy_index = proxy_index + 1
            if proxy_index == len(proxies):
                proxy_index = 0

        final_url = BASE_TARGET_URL4.format(quote_plus(part))
        response = request_by_proxy(final_url)

        soup = BeautifulSoup(response.content, 'html.parser')
        rows = soup.select('.col-lg-9.col-sm-8.col-xs-12 > div')

        # typeName = ""
        for row in rows:
            """
            if row.has_attr("id") and "aaiaPartTypeName" in row.attrs.get("id"):
                typeName = row.select_one('h4').text
                continue
            """
            if row.has_attr("class") and "searchResults" in row.attrs.get("class"):
                productName = row.select_one('span[id$="_lblProductName"]').text
                info = ""
                applicationSummary = ""
                cross = ""
                partNo = ""
                mfrName = ""

                try:
                    info = row.select_one('.searchItems-info h4').text
                except:
                    pass

                try:
                    applicationSummary = row.select_one('.searchItems-info p').text
                except:
                    pass

                try:
                    cross = row.select_one('thead > tr > th').text
                except:
                    pass

                try:
                    partNo = row.select_one('tbody > tr > th').text
                except:
                    pass

                try:
                    mfrName = row.select_one('tbody > tr > td').text
                except:
                    pass

                final_postings.append(
                    [part, productName, info, applicationSummary, cross, partNo, mfrName])
                append_df_to_excel4(
                    output_filename + ".xlsx", [part, productName, info, applicationSummary, cross, partNo, mfrName])

                if models.dorman.objects.filter(original_part_number=part, product_name=productName, part_no=partNo, manufacturer_name=mfrName).exists() is False:
                    models.dorman.objects.create(original_part_number=part, product_name=productName, product_info=info,
                                                 application_summary=applicationSummary, cross=cross, part_no=partNo, manufacturer_name=mfrName)

        if flag_stop == True:
            break

    return final_postings

def new_search4(request):
    output_filename = "static/result/dorman_output"
    if request.POST.get('ckb_autonew') == 'on':
        fileIndex = 1
        fname = output_filename
        while os.path.isfile(fname + ".xlsx") == True:
            fname = "%s_%06d" % (output_filename, fileIndex)
            fileIndex = fileIndex + 1
        output_filename = fname

    global server_status
    final_postings = []

    if request.method == 'POST' and request.FILES:
        input4 = request.FILES['input_file']
        fs = FileSystemStorage("upload_dorman/")
        filename = fs.save(input4.name, input4)
        uploaded_file = fs.path(filename)
        print(uploaded_file)
    else:
        return final_postings
    book = load_workbook(uploaded_file,data_only=True)
    worksheet = book.get_sheet_by_name('Sheet1')
    for i, row_cells in enumerate(worksheet.iter_rows()):
        if i == 0:
            continue
        for cell in row_cells:
            print('%s: cell.value=%s' % (cell, cell.value))
            final_url = BASE_TARGET_URL4.format(quote_plus(str(cell.value)))
            print(final_url)
            response = request_by_proxy(final_url)
            soup = BeautifulSoup(response.content, 'html.parser')
            rows = soup.select('.col-lg-9.col-sm-8.col-xs-12 > div')
            for row in rows:
                if row.has_attr("class") and "searchResults" in row.attrs.get("class"):
                    productName = row.select_one('span[id$="_lblProductName"]').text
                    info = ""
                    applicationSummary = ""
                    cross = ""
                    partNo = ""
                    mfrName = ""
                    try:
                        info = row.select_one('.searchItems-info h4').text
                    except:
                        pass
                    try:
                        applicationSummary = row.select_one('.searchItems-info p').text
                    except:
                        pass
                    try:
                        cross = row.select_one('thead > tr > th').text
                    except:
                        pass
                    try:
                        partNo = row.select_one('tbody > tr > th').text
                    except:
                        pass
                    try:
                        mfrName = row.select_one('tbody > tr > td').text
                    except:
                        pass
                    final_postings.append([cell.value, productName, info, applicationSummary, cross, partNo, mfrName])
                    append_df_to_excel4_new(output_filename + ".xlsx", [cell.value, productName, info, applicationSummary, cross, partNo, mfrName])
                    
                    if models.dorman.objects.filter(original_part_number=cell.value, product_name=productName, part_no=partNo, manufacturer_name=mfrName).exists() is False:
                        models.dorman.objects.create(original_part_number=cell.value, product_name=productName, product_info=info,application_summary=applicationSummary, cross=cross, part_no=partNo, manufacturer_name=mfrName)
            if flag_stop == True:
                break
    fs.delete(uploaded_file)
    return final_postings

def search5(request):
    """
        login
    """
    session = requests.Session()
    session.mount('http://', HTTPAdapter(max_retries=1))
    session.mount('https://', HTTPAdapter(max_retries=1))

    login_data = {
        'resolution': '',
        'nexpartuname': username,
        'pwd': password,
        'B1': 'Login',
        'part': '',
        'linecode': '',
        'qty': '',
        'product': '2'
    }

    global server_status
    response = session.post(f'{BASE_TARGET_URL5}/do_login.php', headers=headers5, data=login_data)
    if '?err=1' in response.url:
        server_status = server_status + "\n" + "---------- Nexcat Login error! ----------"
        print(server_status)
        return render(request, 'base.html')

    """
        Scrape
    """
    output_filename = "static/result/nexcat_output"
    if request.POST.get('ckb_autonew') == 'on':
        fileIndex = 1
        fname = output_filename
        while os.path.isfile(fname + ".xlsx") == True:
            fname = "%s_%06d" % (output_filename, fileIndex)
            fileIndex = fileIndex + 1
        output_filename = fname

    search = request.POST.get('search')
    part_numbers = search.split()

    final_postings = []
    for part in part_numbers:
        final_url = f'{BASE_TARGET_URL5}/smartpage_redirect.php?isModal=1&mfrlinecode=MVC&partnumber={quote_plus(part)}&acesptermid=7536&app=nexpart&group=&usg=&packqty=&selldesc=&listprice=&dealerprice=&regularprice=&yousaveprice=&privateline=MEVOTECH%20CONTROL%20ARMS&showsellprice=1&yousave=1&addldesc='
        response = session.get(final_url, headers=headers5, verify=False)
        print(response.status_code, response.url)
        server_status = server_status + "\n" + "{}, {}".format(response.status_code, response.url)

        html = fromstring(response.content)

        for row in html.xpath('//div[@id="oeInterchangeBtnContent"]/div[@class="pairWrap"]'):
            parts = row.xpath('.//div/text()')

            final_postings.append([part, parts[1], '', parts[0]])
            append_df_to_excel5(output_filename + ".xlsx", [part, parts[1], '', parts[0]])

            if models.nexcat.objects.filter(original_part_number=part, oe_part_no=parts[1], manufacturer_name=parts[0]).exists() is False:
                models.nexcat.objects.create(
                    original_part_number=part, oe_part_no=parts[1], am_part_no='', manufacturer_name=parts[0])

        for row in html.xpath('//div[@id="amInterchangeBtnContent"]/div[@class="pairWrap"]'):
            parts = row.xpath('.//div/text()')

            final_postings.append([part, '', parts[1], parts[0]])
            append_df_to_excel5(output_filename + ".xlsx", [part, '', parts[1], parts[0]])

            if models.nexcat.objects.filter(original_part_number=part, am_part_no=parts[1], manufacturer_name=parts[0]).exists() is False:
                models.nexcat.objects.create(
                    original_part_number=part, oe_part_no='', am_part_no=parts[1], manufacturer_name=parts[0])

        if flag_stop == True:
            break

    return final_postings


def scrapesearch(request):
    global server_status
    final_postings = []
    output_filename = "static/result/dorman_output"
    if request.POST.get('ckb_autonew') == 'on':
        fileIndex = 1
        fname = output_filename
        while os.path.isfile(fname + ".xlsx") == True:
            fname = "%s_%06d" % (output_filename, fileIndex)
            fileIndex = fileIndex + 1
        output_filename = fname
    if request.method == 'POST' and request.FILES:
        input4 = request.FILES['input_file']
        fs = FileSystemStorage("upload_dorman/")
        filename = fs.save(input4.name, input4)
        uploaded_file = fs.path(filename)
        df=pd.read_excel(uploaded_file,usecols="A")
        codelist=df['Compressed Old number'].to_list()
        event = codelist
        fp = freeproxy.WriteProxyList()
        fp.fetchandwriteproxy('/home/ubuntu/work/DDS_Web/proxyurl.lst')
        scrape(event,output=output_filename)
        #print("scrape-------------",final)
        #final_postings.append(final)
    #return final_postings


def scrape(event={},context={},output=None):
    finalist=[]
    for i in event:
        finalist=crawl(i,output)
    return finalist
    #for i in finalist:
     #   return i


proxy_index = 0
part_index = 0

def request_by_proxy(url):
    global proxies
    global proxy_index
    global part_index
    global server_status

    proxy = proxies[proxy_index]

    try:
        response = requests.get(url, proxies={"https": "https://" + proxy},
                                headers=headers4[random.randint(0, len(headers4) - 1)], timeout=3)
        if response.url != url:
            raise Exception("Captcha Error")
        print(part_index, proxy_index, proxy, response.status_code, response.url)
        server_status = server_status + "\n" + "{}, {}".format(response.status_code, response.url)
    except:
        if part_index < 5:
            proxies.remove(proxy)
            if proxy_index > 0:
                proxy_index = proxy_index - 1
        else:
            proxy_index = proxy_index + 1

        if proxy_index == len(proxies):
            proxy_index = 0
        if len(proxies) < 10:
            proxies = copy.deepcopy(proxy_list)

        part_index = 1
        response = request_by_proxy(url)

    return response


def new_search5(request):
    output_filename = "static/result/nexcat_output"
    if request.POST.get('ckb_autonew') == 'on':
        fileIndex = 1
        fname = output_filename
        while os.path.isfile(fname + ".xlsx") == True:
            fname = "%s_%06d" % (output_filename, fileIndex)
            fileIndex = fileIndex + 1
        output_filename = fname

    global server_status
    final_postings = []

    if request.method == 'POST' and request.FILES:
        input5 = request.FILES['input_file']
        fs = FileSystemStorage()
        filename = fs.save("upload/nexcat_input.xlsx", input5)
        uploaded_file = fs.path(filename)
    else:
        return final_postings

    book = load_workbook(uploaded_file, data_only=True)
    sheet = book.active
    rows = sheet.rows

    if platform == "linux" or platform == "linux2":
        display = Display(visible=0, size=(1024, 1000))
        display.start()

    chrome_options = Options()
    chrome_options.add_argument("window-size=1024,1000")
    chrome_options.add_argument("--disable-notifications")
    chrome_options.add_argument("--no-sandbox")
    user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.149 Safari/537.36'
    chrome_options.add_argument(f'user-agent={user_agent}')
    chrome_options.add_argument('--verbose')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--disable-software-rasterizer')

    # if platform == "win32" or platform == "win64":
    #     chrome_options.add_argument("--headless")

    driver = webdriver.Chrome(options=chrome_options)

    nexcat_login(driver)
    driver.get(f'{BASE_TARGET_URL5}/nexpartbg.php?newHilite=TT02')

    for row in rows:
        time.sleep(3)
        try:
            driver.find_element_by_xpath('//*[@id="smpgDiv_iconClose"]').click()
            time.sleep(1)
        except:
            pass
        try:
            driver.find_element_by_xpath(
                '//div[contains(@class, "npvue-button2--secondary")]').click()
            time.sleep(1)
        except:
            pass

        part = []
        for cell in row:
            part.append(str(cell.value).strip())

        server_status = server_status + "\n" + part[0] + " " + part[1]
        print(server_status)

        # part_input = driver.find_element_by_xpath('//input[@class="npvue-input2__input"]')
        part_input = driver.find_element_by_xpath('//input[contains(@id, "input2-")]')
        part_input.click()
        part_input.clear()
        part_input.send_keys(part[0])
        time.sleep(1)
        part_input.send_keys(Keys.RETURN)
        time.sleep(3)

        try:
            brand_select = driver.find_element_by_xpath(
                '//div[@class="pli__partParam_brand"]//span[contains(@class, "select2-container")]')
            if not brand_select.is_displayed() or not brand_select.is_enabled():
                raise Exception()
        except:
            continue

        brand_select.click()
        time.sleep(1)

        rows = driver.find_elements_by_xpath('//li[@class="select2-results__option"]')
        for row in rows:
            if row.text.strip() == part[1]:
                row.click()
                time.sleep(3)
                break

        try:
            smartPageLink = driver.find_element_by_xpath('//span[@class="smartPageLink"]')
        except:
            continue

        smartPageLink.click()
        time.sleep(3)

        title = driver.find_element_by_xpath('//span[contains(@class, "sp_title")]').text.strip()

        rows = driver.find_elements_by_xpath(
            '//div[@id="oeInterchangeBtnContent"]/div[@class="pairWrap"]')
        for row in rows:
            manufacturer = row.find_element_by_xpath('./div[@class="col4_1"]').text.strip()
            amPartNumber = row.find_element_by_xpath('./div[@class="col4_2"]').text.strip()

            final_postings.append([part[0], part[1], title, manufacturer, amPartNumber])
            append_df_to_excel5_new(output_filename + ".xlsx",
                                    [part[0], part[1], title, manufacturer, amPartNumber])

            if models.nexcat_new.objects.filter(part_number=part[0], brand=part[1], title=title, manufacturer=manufacturer, am_part_no=amPartNumber).exists() is False:
                models.nexcat_new.objects.create(
                    part_number=part[0], brand=part[1], title=title, manufacturer=manufacturer, am_part_no=amPartNumber)

        driver.find_element_by_xpath('//*[@id="amInterchangeBtn"]').click()
        time.sleep(1)
        rows = driver.find_elements_by_xpath(
            '//div[@id="amInterchangeBtnContent"]/div[@class="pairWrap"]')
        for row in rows:
            manufacturer = row.find_element_by_xpath('./div[@class="col4_1"]').text.strip()
            amPartNumber = row.find_element_by_xpath('./div[@class="col4_2"]').text.strip()

            final_postings.append([part[0], part[1], title, manufacturer, amPartNumber])
            append_df_to_excel5_new(output_filename + ".xlsx",
                                    [part[0], part[1], title, manufacturer, amPartNumber])

            if models.nexcat_new.objects.filter(part_number=part[0], brand=part[1], title=title, manufacturer=manufacturer, am_part_no=amPartNumber).exists() is False:
                models.nexcat_new.objects.create(
                    part_number=part[0], brand=part[1], title=title, manufacturer=manufacturer, am_part_no=amPartNumber)

        if flag_stop == True:
            break

    fs.delete(uploaded_file)

    driver.close()
    driver.quit()

    if platform == "linux" or platform == "linux2":
        display.stop()

    return final_postings


def nexcat_login(driver: webdriver):
    global server_status

    driver.get(BASE_TARGET_URL5)
    try:
        time.sleep(3)
        login = driver.find_element_by_xpath('//*[@id="username"]')
        if not login.is_displayed() or not login.is_enabled():
            raise Exception()
    except Exception:
        server_status = server_status + "\nAlready logged in!"
        print("Already logged in!")
        return True

    driver.find_element_by_xpath('//label[@class="rememeber"]').click()
    time.sleep(1)

    login_username = driver.find_element_by_xpath('//*[@id="username"]')
    login_username.click()
    login_username.clear()
    login_username.send_keys(username)
    time.sleep(1)
    login_password = driver.find_element_by_xpath('//*[@id="password"]')
    login_password.click()
    login_password.clear()
    login_password.send_keys(password)
    time.sleep(1)

    driver.find_element_by_xpath('//input[@type="submit" and @value="Login"]').click()
    time.sleep(3)

    try:
        login = driver.find_element_by_xpath('//*[@id="username"]')
        if not login.is_displayed() or not login.is_enabled():
            raise Exception()
    except Exception:
        server_status = server_status + "\nLogin Successful!"
        print("Login Successful!")
        return True

    return False


def search6(request):
    output_filename = "static/result/trw_output"
    if request.POST.get('ckb_autonew') == 'on':
        fileIndex = 1
        fname = output_filename
        while os.path.isfile(fname + ".xlsx") == True:
            fname = "%s_%06d" % (output_filename, fileIndex)
            fileIndex = fileIndex + 1
        output_filename = fname

    global server_status
    final_postings = []

    if request.method == 'POST' and request.FILES:
        input6 = request.FILES['input_file']
        fs = FileSystemStorage()
        filename = fs.save("upload/trw_input.xlsx", input6)
        uploaded_file = fs.path(filename)
    else:
        return final_postings

    book = load_workbook(uploaded_file, data_only=True)
    sheet = book.active
    rows = sheet.rows

    if platform == "linux" or platform == "linux2":
        display = Display(visible=0, size=(1024, 1000))
        display.start()

    chrome_options = Options()
    chrome_options.add_argument("window-size=1024,1000")
    chrome_options.add_argument("--disable-notifications")
    chrome_options.add_argument("--no-sandbox")
    user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.149 Safari/537.36'
    chrome_options.add_argument(f'user-agent={user_agent}')
    chrome_options.add_argument('--verbose')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--disable-software-rasterizer')

    # if platform == "win32" or platform == "win64":
    #     chrome_options.add_argument("--headless")

    driver = webdriver.Chrome(options=chrome_options)

    try:
        driver.get(BASE_TARGET_URL6)
    except:
        server_status = server_status + "\n" + "---------- No connect to showmetheparts.com -----"
        print(server_status)

        driver.close()
        driver.quit()

        if platform == "linux" or platform == "linux2":
            display.stop()

        return render(request, 'base.html')

    for row in rows:
        part = []
        try:
            year = int(row[0].value)
            if year < 1930 or year > 2050:
                continue
        except:
            continue

        for cell in row:
            part.append(cell.value)

        try:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located(
                (By.CSS_SELECTOR, "#boundlist-1327-listEl > ul > li")))
            year = driver.find_element_by_id("combo-1017-inputEl")
            year.click()
            year.send_keys(part[0])
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
                (By.CSS_SELECTOR, "#boundlist-1327-listEl > ul > li")))
            time.sleep(1)
            year.send_keys(Keys.RETURN)
        except TimeoutException:
            server_status = server_status + "\n" + "---------- Timeout while locating 'Year List' ----------"
            print(server_status)
        except:  # NoSuchElementException
            server_status = server_status + "\n" + "---------- Not able to locate 'Year List' ----------"
            print(server_status)

        try:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located(
                (By.CSS_SELECTOR, "#boundlist-1329-listEl > ul > li")))
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
                (By.CSS_SELECTOR, "#boundlist-1329-listEl > ul > li")))
            make = driver.find_element_by_id("combo-1019-inputEl")
            make.send_keys(part[1].upper())
            time.sleep(1)
            make.send_keys(Keys.RETURN)
        except TimeoutException:
            server_status = server_status + "\n" + "---------- Timeout while locating 'Make List' ----------"
            print(server_status)
        except:  # NoSuchElementException
            server_status = server_status + "\n" + "---------- Not able to locate 'Make List' ----------"
            print(server_status)

        try:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located(
                (By.CSS_SELECTOR, "#boundlist-1331-listEl > ul > li")))
            model = driver.find_element_by_id("combo-1021-inputEl")
            model.send_keys(part[2].upper())
            time.sleep(1)
            model.send_keys(Keys.RETURN)
        except TimeoutException:
            server_status = server_status + "\n" + "---------- Timeout while locating 'Model List' ----------"
            print(server_status)
        except:  # NoSuchElementException
            server_status = server_status + "\n" + "---------- Not able to locate 'Model List' ----------"
            print(server_status)

        try:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located(
                (By.CSS_SELECTOR, "#boundlist-1333-listEl > ul > li")))
            product = driver.find_element_by_id("combo-1023-inputEl")
            product.send_keys(part[3].upper())
            time.sleep(1)
            product.send_keys(Keys.RETURN)
        except TimeoutException:
            try:
                WebDriverWait(driver, 3).until(EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "#boundlist-1386-listEl > ul > li")))
                product = driver.find_element_by_id("combo-1023-inputEl")
                product.send_keys(part[3].upper())
                time.sleep(1)
                product.send_keys(Keys.RETURN)
            except TimeoutException:
                server_status = server_status + "\n" + "---------- Timeout while locating 'Product List' ----------"
                print(server_status)
        except:  # NoSuchElementException
            server_status = server_status + "\n" + "---------- Not able to locate 'Product List' ----------"
            print(server_status)

        try:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located(
                (By.CSS_SELECTOR, "#gridview-1042-body > tr")))
            time.sleep(1)
            tbl = driver.find_element_by_id("gridview-1042-body")
            rows = driver.execute_script("""
                return arguments[0].children
            """, tbl)
        except TimeoutException:
            server_status = server_status + "\n" + "---------- Timeout while locating 'Parts Table' ----------"
            print(server_status)
        except:  # NoSuchElementException
            server_status = server_status + "\n" + "---------- Not able to locate 'Parts Table' ----------"
            print(server_status)

        try:
            for row in rows:
                partNumber = row.find_element_by_css_selector('.PartNumberLink').text
                partType = row.find_element_by_css_selector(
                    '.x-grid-cell-headerId-colAppPartType div').text
                # brand = row.find_element_by_css_selector('.x-grid-cell-headerId-colAppBrand div').text
                # comment = row.find_element_by_css_selector('.x-grid-cell-headerId-colAppComment div').text
                application = row.find_element_by_css_selector(
                    '.x-grid-cell-headerId-colAppApplication div').text

                final_postings.append([part[0], part[1], part[2], part[3],
                                       partNumber, partType, application])
                append_df_to_excel6(
                    output_filename + ".xlsx", [part[0], part[1], part[2], part[3], partNumber, partType, application])

                if models.trw.objects.filter(year=part[0], make=part[1], model=part[2], product=part[3], part_number=partNumber, application=application).exists() is False:
                    models.trw.objects.create(year=part[0], make=part[1], model=part[2], product=part[3],
                                              part_number=partNumber, part_type=partType, application=application)

        except:
            server_status = server_status + "\n" + "---------- Parts Table Rows Error() ----------"
            print(server_status)

        if flag_stop == True:
            break

    fs.delete(uploaded_file)

    driver.close()
    driver.quit()

    if platform == "linux" or platform == "linux2":
        display.stop()

    return final_postings


def export_data(final_postings):
    global server_status

    items = []
    for site_postings in final_postings:
        for post in site_postings:
            items.append({
                'Make': post[0],
                'Model': post[1],
                'Product': post[2],
                'Part Number': post[3],
                'Part Type': post[4],
                'Application': post[5]
            })

    try:
        df = pd.DataFrame(items)
        df.to_excel("output6.xlsx", index=False)
    except:
        server_status = server_status + "\n" + "---------- Error export data into Excel ----------"
        print(server_status)


def append_df_to_excel(filename, data):
    try:
        book = load_workbook(filename)
        sheet = book.active
    except:
        book = Workbook()
        sheet = book.active
        headers = ['original part number', 'partNo', 'Manufacturer Name', 'Product Type']
        sheet.append(headers)

    sheet.append(data)
    book.save(filename)


def append_df_to_excel4(filename, data):
    try:
        book = load_workbook(filename)
        sheet = book.active
    except:
        book = Workbook()
        sheet = book.active
        headers = ['original part number', 'Product Name', 'Product Info',
                   'Application Summary', 'Cross', 'Part No', 'Mfr Name']
        sheet.append(headers)

    sheet.append(data)
    book.save(filename)

def append_df_to_excel4_new(filename, data):
    try:
        book = load_workbook(filename)
        sheet = book.active
    except:
        book = Workbook()
        sheet = book.active
        headers = ['productName', 'info', 'applicationSummary', 'cross', 'partNo', 'mfrName','Oedetails']
        sheet.append(headers)

    sheet.append(data)
    book.save(filename)

def append_df_to_excel5(filename, data):
    try:
        book = load_workbook(filename)
        sheet = book.active
    except:
        book = Workbook()
        sheet = book.active
        headers = ['original part number', 'OE Part Number', 'AM Part Number', 'Manufacturer']
        sheet.append(headers)

    sheet.append(data)
    book.save(filename)


def append_df_to_excel5_new(filename, data):
    try:
        book = load_workbook(filename)
        sheet = book.active
    except:
        book = Workbook()
        sheet = book.active
        headers = ['part number', 'Brand', 'Title', 'Manufacturer', 'AM Part Number']
        sheet.append(headers)

    sheet.append(data)
    book.save(filename)


def append_df_to_excel6(filename, data):
    try:
        book = load_workbook(filename)
        sheet = book.active
    except:
        book = Workbook()
        sheet = book.active
        headers = ['Year', 'Make', 'Model', 'Product', 'Part Number', 'Part Type', 'Application']
        sheet.append(headers)

    sheet.append(data)
    book.save(filename)


def search_db1(request):
    search = request.POST.get('search')
    part_numbers = search.split()

    final_postings = []
    for part in part_numbers:
        rows = models.fme_cat.objects.filter(original_part_number=part)

        for row in rows:
            final_postings.append([row.original_part_number, row.part_no,
                                   row.manufacturer_name, row.product_type])

        if flag_stop == True:
            break

    return final_postings


def search_db2(request):
    search = request.POST.get('search')
    part_numbers = search.split()

    final_postings = []
    for part in part_numbers:
        rows = models.fme_mx.objects.filter(original_part_number=part)

        for row in rows:
            final_postings.append([row.original_part_number, row.part_no,
                                   row.manufacturer_name, row.product_type])

        if flag_stop == True:
            break

    return final_postings


def search_db3(request):
    search = request.POST.get('search')
    part_numbers = search.split()

    final_postings = []
    for part in part_numbers:
        rows = models.mevotech.objects.filter(original_part_number=part)

        for row in rows:
            final_postings.append([row.original_part_number, row.part_no,
                                   row.manufacturer_name, row.product_type])

        if flag_stop == True:
            break

    return final_postings


def search_db4(request):
    search = request.POST.get('search')
    part_numbers = search.split()

    final_postings = []
    for part in part_numbers:
        rows = models.dorman.objects.filter(original_part_number=part)

        for row in rows:
            final_postings.append([row.original_part_number, row.product_name, row.product_info,
                                   row.application_summary, row.cross, row.part_no, row.manufacturer_name])

        if flag_stop == True:
            break

    return final_postings


def search_db5(request):
    search = request.POST.get('search')
    part_numbers = search.split()

    final_postings = []
    for part in part_numbers:
        rows = models.nexcat.objects.filter(original_part_number=part)

        for row in rows:
            final_postings.append([row.original_part_number, row.oe_part_no,
                                   row.am_part_no, row.manufacturer_name])

        if flag_stop == True:
            break

    return final_postings


def new_search_db5(request):
    part_numbers = []
    if request.method == 'POST' and request.FILES:
        input5 = request.FILES['input_file']
        fs = FileSystemStorage()
        filename = fs.save("upload/nexcat_input.xlsx", input5)
        uploaded_file = fs.path(filename)

        book = load_workbook(uploaded_file, data_only=True)
        sheet = book.active
        rows = sheet.rows

        for row in rows:
            values = []
            for cell in row:
                values.append(cell.value)

            part_numbers.append(values)

        fs.delete(uploaded_file)
    else:
        return []

    final_postings = []
    for part in part_numbers:
        rows = models.nexcat_new.objects.filter(part_number=part[0], brand=part[1])

        for row in rows:
            final_postings.append([row.part_number, row.brand, row.title,
                                   row.manufacturer, row.am_part_no])
        if flag_stop == True:
            break

    return final_postings


def search_db6(request):
    part_numbers = []
    if request.method == 'POST' and request.FILES:
        input6 = request.FILES['input_file']
        fs = FileSystemStorage()
        filename = fs.save("upload/trw_input.xlsx", input6)
        uploaded_file = fs.path(filename)

        book = load_workbook(uploaded_file, data_only=True)
        sheet = book.active
        rows = sheet.rows

        for row in rows:
            values = []
            try:
                year = int(row[0].value)
                if year < 1930 or year > 2050:
                    continue
            except:
                continue

            for cell in row:
                values.append(cell.value)

            part_numbers.append(values)

        fs.delete(uploaded_file)
    else:
        return []

    final_postings = []
    for part in part_numbers:
        rows = models.trw.objects.filter(year=part[0], make=part[1], model=part[2], product=part[3])

        for row in rows:
            final_postings.append([row.year, row.make, row.model, row.product,
                                   row.part_number, row.part_type, row.application])
        if flag_stop == True:
            break

    return final_postings

